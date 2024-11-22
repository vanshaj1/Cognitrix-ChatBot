# # transcripts/views.py

# from django.shortcuts import render
# from youtube_transcript_api import YouTubeTranscriptApi
# from django.http import JsonResponse
# import re
# import os
# from openpyxl import Workbook, load_workbook
# import requests

# # Add your YouTube Data API key here
# YOUTUBE_API_KEY = 'AIzaSyCW6H06Tz-Hjph65jtk4WoUXF4u-RXrH6E'

# # Helper function to extract video ID or playlist ID
# def extract_video_or_playlist_id(youtube_url):
#     video_regex = r"(?<=v=)[^&]+"
#     playlist_regex = r"(?<=list=)[^&]+"

#     video_match = re.search(video_regex, youtube_url)
#     playlist_match = re.search(playlist_regex, youtube_url)

#     if playlist_match:
#         return {'playlist_id': playlist_match.group(0)}
#     elif video_match:
#         return {'video_id': video_match.group(0)}
#     return None

# # Function to save transcript to Excel, including the video link and combined column
# def save_transcript_to_excel(video_id, transcript_text, video_link):
#     file_path = 'transcripts/transcripts.xlsx'

#     # Create or load the workbook
#     if os.path.exists(file_path):
#         workbook = load_workbook(file_path)
#         sheet = workbook.active
#     else:
#         workbook = Workbook()
#         sheet = workbook.active
#         # Add headers for new sheet
#         sheet.append(['Video ID', 'Transcript', 'Video Link', 'Video Link + Transcript'])

#     # Format the combined value with line breaks
#     combined_value = f"Link: {video_link}\nTranscript: {transcript_text}"

#     # Append the row data to the sheet
#     sheet.append([video_id, transcript_text, video_link, combined_value])
#     workbook.save(file_path)

# # Function to get all video IDs from a YouTube playlist
# def get_video_ids_from_playlist(playlist_id):
#     base_url = 'https://www.googleapis.com/youtube/v3/playlistItems'
#     params = {
#         'part': 'contentDetails',
#         'playlistId': playlist_id,
#         'maxResults': 50,  # Maximum results per request
#         'key': YOUTUBE_API_KEY
#     }

#     video_ids = []
#     next_page_token = None

#     while True:
#         if next_page_token:
#             params['pageToken'] = next_page_token

#         response = requests.get(base_url, params=params)
#         data = response.json()

#         if 'items' in data:
#             video_ids += [item['contentDetails']['videoId'] for item in data['items']]

#         # Check for next page token (to get more videos if available)
#         next_page_token = data.get('nextPageToken')
#         if not next_page_token:
#             break

#     return video_ids

# def get_transcript(request):
#     if request.method == 'POST':
#         youtube_link = request.POST.get('youtubeLink')
#         extracted_ids = extract_video_or_playlist_id(youtube_link)

#         if not extracted_ids:
#             return JsonResponse({'error': 'Invalid YouTube URL'}, status=400)

#         try:
#             if 'playlist_id' in extracted_ids:
#                 # It's a playlist URL, fetch all video IDs
#                 playlist_id = extracted_ids['playlist_id']
#                 video_ids = get_video_ids_from_playlist(playlist_id)

#                 for video_id in video_ids:
#                     try:
#                         transcript = YouTubeTranscriptApi.get_transcript(video_id)
#                         transcript_text = " ".join([entry['text'] for entry in transcript])
#                         video_url = f"https://www.youtube.com/watch?v={video_id}"
#                         save_transcript_to_excel(video_id, transcript_text, video_url)
#                     except Exception as e:
#                         # Skip videos that don't have transcripts
#                         print(f"Error getting transcript for video {video_id}: {str(e)}")
                
#                 return JsonResponse({'message': f'Transcripts for {len(video_ids)} videos saved successfully!'})

#             elif 'video_id' in extracted_ids:
#                 # It's a single video URL
#                 video_id = extracted_ids['video_id']
#                 transcript = YouTubeTranscriptApi.get_transcript(video_id)
#                 transcript_text = " ".join([entry['text'] for entry in transcript])
#                 video_url = f"https://www.youtube.com/watch?v={video_id}"
#                 save_transcript_to_excel(video_id, transcript_text, video_url)
#                 return JsonResponse({'message': 'Transcript successfully saved!'})

#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

#     return JsonResponse({'error': 'Invalid request method'}, status=400)

# def index(request):
#     return render(request, 'transcripts/index.html')

# transcripts/views.py

from django.shortcuts import render
from youtube_transcript_api import YouTubeTranscriptApi
from django.http import JsonResponse
import re
import os
from openpyxl import Workbook, load_workbook
import requests

# Add your YouTube Data API key here
YOUTUBE_API_KEY = 'AIzaSyCW6H06Tz-Hjph65jtk4WoUXF4u-RXrH6E'

# Helper function to extract video ID or playlist ID
def extract_video_or_playlist_id(youtube_url):
    video_regex = r"(?<=v=)[^&]+"
    playlist_regex = r"(?<=list=)[^&]+"

    video_match = re.search(video_regex, youtube_url)
    playlist_match = re.search(playlist_regex, youtube_url)

    if playlist_match:
        return {'playlist_id': playlist_match.group(0)}
    elif video_match:
        return {'video_id': video_match.group(0)}
    return None

# Function to save transcript to Excel
def save_transcript_to_excel(video_id, transcript_text, video_link, playlist_name):
    file_path = 'transcripts/transcripts.xlsx'

    # Create or load the workbook
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # Add headers for new sheet
        sheet.append(['Playlist Name', 'Video ID', 'Transcript', 'Video Link', 'Video Link + Transcript'])

    # Format the combined value with line breaks
    combined_value = f"Link: {video_link}\nTranscript: {transcript_text}"

    # Append the row data to the sheet
    sheet.append([playlist_name, video_id, transcript_text, video_link, combined_value])
    workbook.save(file_path)

# Function to get all video IDs from a YouTube playlist
def get_video_ids_from_playlist(playlist_id):
    base_url = 'https://www.googleapis.com/youtube/v3/playlistItems'
    params = {
        'part': 'contentDetails',
        'playlistId': playlist_id,
        'maxResults': 50,
        'key': YOUTUBE_API_KEY
    }

    video_ids = []
    next_page_token = None

    while True:
        if next_page_token:
            params['pageToken'] = next_page_token

        response = requests.get(base_url, params=params)
        data = response.json()

        if 'items' in data:
            video_ids += [item['contentDetails']['videoId'] for item in data['items']]

        next_page_token = data.get('nextPageToken')
        if not next_page_token:
            break

    return video_ids

def get_transcript(request):
    if request.method == 'POST':
        youtube_link = request.POST.get('youtubeLink')
        extracted_ids = extract_video_or_playlist_id(youtube_link)

        if not extracted_ids:
            return JsonResponse({'error': 'Invalid YouTube URL'}, status=400)

        try:
            if 'playlist_id' in extracted_ids:
                # It's a playlist URL
                playlist_id = extracted_ids['playlist_id']
                playlist_name = youtube_link.split('=')[1]  # Extract playlist name from URL
                video_ids = get_video_ids_from_playlist(playlist_id)

                for video_id in video_ids:
                    try:
                        transcript = YouTubeTranscriptApi.get_transcript(video_id)
                        transcript_text = " ".join([entry['text'] for entry in transcript])
                        video_url = f"https://www.youtube.com/watch?v={video_id}"
                        save_transcript_to_excel(video_id, transcript_text, video_url, playlist_name)
                    except Exception as e:
                        print(f"Error getting transcript for video {video_id}: {str(e)}")
                
                return JsonResponse({'message': f'Transcripts for {len(video_ids)} videos saved successfully!'})

            elif 'video_id' in extracted_ids:
                # It's a single video URL
                video_id = extracted_ids['video_id']
                transcript = YouTubeTranscriptApi.get_transcript(video_id)
                transcript_text = " ".join([entry['text'] for entry in transcript])
                video_url = f"https://www.youtube.com/watch?v={video_id}"
                save_transcript_to_excel(video_id, transcript_text, video_url, "")
                return JsonResponse({'message': 'Transcript successfully saved!'})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=400)

def remove_playlist(request, playlist_name):
    file_path = 'transcripts/transcripts.xlsx'

    if not os.path.exists(file_path):
        return JsonResponse({'error': 'Excel file does not exist'}, status=404)

    workbook = load_workbook(file_path)
    sheet = workbook.active

    rows_to_keep = []
    for row in sheet.iter_rows(values_only=True):
        if row[0] != playlist_name:  # Keep rows that don't match the playlist name
            rows_to_keep.append(row)

    # Clear existing sheet and write back the rows to keep
    sheet.delete_rows(1, sheet.max_row)
    for row in rows_to_keep:
        sheet.append(row)

    workbook.save(file_path)
    return JsonResponse({'message': f'Playlist "{playlist_name}" removed successfully!'})

def index(request):
    # Load the playlists to display on the homepage
    playlists = []
    file_path = 'transcripts/transcripts.xlsx'
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            playlists.append(row[0])  # Append the playlist name

    return render(request, 'transcripts/index.html', {'playlists': set(playlists)})  # Return unique playlist names
